from celery import shared_task
from django.conf import settings
import openpyxl
import os


@shared_task
def ingest_customer_data():
    file_path = os.path.join(settings.BASE_DIR, 'customer_data.xlsx')
    
    # Import here to avoid circular imports
    from .models import Customer
    
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip empty rows
        if not row[0]:
            continue

        customer_id  = row[0]
        first_name   = row[1]
        last_name    = row[2]
        phone_number = row[3]
        monthly_salary = row[4]
        approved_limit = row[5]
        current_debt   = row[6] if row[6] else 0

        Customer.objects.update_or_create(
            customer_id=customer_id,
            defaults={
                'first_name': first_name,
                'last_name': last_name,
                'phone_number': phone_number,
                'monthly_salary': monthly_salary,
                'approved_limit': approved_limit,
                'current_debt': current_debt,
            }
        )
        count += 1

    return f"Successfully ingested {count} customers"


@shared_task
def ingest_loan_data():
    file_path = os.path.join(settings.BASE_DIR, 'loan_data.xlsx')

    from .models import Customer, Loan

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip empty rows
        if not row[0]:
            continue

        customer_id    = row[0]
        loan_id        = row[1]
        loan_amount    = row[2]
        tenure         = row[3]
        interest_rate  = row[4]
        monthly_repayment = row[5]
        emis_paid_on_time = row[6] if row[6] else 0
        start_date     = row[7]
        end_date       = row[8]

        try:
            customer = Customer.objects.get(customer_id=customer_id)
        except Customer.DoesNotExist:
            continue

        Loan.objects.update_or_create(
            loan_id=loan_id,
            defaults={
                'customer': customer,
                'loan_amount': loan_amount,
                'tenure': tenure,
                'interest_rate': interest_rate,
                'monthly_repayment': monthly_repayment,
                'emis_paid_on_time': emis_paid_on_time,
                'start_date': start_date,
                'end_date': end_date,
            }
        )
        count += 1

    return f"Successfully ingested {count} loans"